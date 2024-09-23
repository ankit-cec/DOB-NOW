import os
import openpyxl
from selenium import webdriver
driver = webdriver.Chrome(r"C:\Users\cecmi\Downloads\dist-20240923T162942Z-001\dist\chromedriver.exe")
from time import sleep
import ast
import shutil
from PyPDF2 import PdfReader, PdfWriter
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

if os.path.exists('Signs'):
    # Delete the folder and all of its contents
    shutil.rmtree('Signs')
if not os.path.exists('Signs'):
    os.mkdir('Signs')
if not os.path.exists('Renamed Signs'):
    os.mkdir('Renamed Signs')

source_folder = r"C:\Users\cecmi\Downloads\dist-20240923T162942Z-001\dist\Signs"
destination_folder = r"C:\Users\cecmi\Downloads\dist-20240923T162942Z-001\dist\Renamed Signs"
wb = openpyxl.load_workbook(r"C:\Users\cecmi\Downloads\dist-20240923T162942Z-001\dist\Signdownload_2023.xlsx")
ws = wb['Sheet1']
links = [a.value for a in ws['d']][1:]
names = [a.value for a in ws['b']][1:]
wb.save('Signdownload_2023.xlsx')
print(len(links))
options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-automation'])
options.add_argument('--disable-blink-features=AutomationControlled')
# options.add_argument('--headless')
options.add_experimental_option('prefs', {
    # 'download.default_directory': r'C:\Users\Doree\OneDrive\Desktop\dist\Signs',
    'download.default_directory': r"C:\Users\cecmi\Downloads\dist-20240923T162942Z-001\dist\Signs Download",
    # 'download.default_directory': current_directory,
    'download.prompt_for_download': False,
    "download.directory_upgrade": True,
    'plugins.always_open_pdf_externally': True})
driver = webdriver.Chrome(options=options)
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
    'source': """
           Object.defineProperty(navigator,'webdriver',{
                get:()=>undefined
           })
    """
})
driver.maximize_window()
sleep(1)
# for li in range(1118,len(links)):
for li in range(0,len(links)):
    # if i==3:
    #     continue
    link = links[li]

    try:
        link = ast.literal_eval(link) if isinstance(link, str) else link
    except (ValueError, SystemError):
        print(f"Skipping malformed link at index {li}: {link}")
        continue


    name = names[li].replace('/', '').replace('\\', '').replace('<', '').replace('>', '').replace('*', '').replace('|',
                                                                                                                   '').replace(
        '?', '').replace(':', '').replace('"', '')
    url = ''
    for l in range(len(link)):
        if 'http' not in link[l]:
            continue
        url = link[l]
        # break
        # if url == '':
        #     continue
        print(f'{li}_{l}', url)
        driver.get(url)
        sleep(2)
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="certCanvas"]/div[1]/button')))
        sleep(2)
        driver.find_element(By.XPATH, '//*[@id="certCanvas"]/div[1]/button').click()
        sleep(2)
        pdf_files = []
        tt = 0
        for t in range(60):
            tt = t
            # Get all PDF files in the source folder
            pdf_files = [f for f in os.listdir(source_folder) if f.endswith('.pdf')]
            if len(pdf_files) != 0:
                break
            else:
                sleep(1)
        if tt == 59:
            continue
        # print(len(pdf_files))
        # Work Through All pdf Signs
        sleep(1)
        for i, file_name in enumerate(pdf_files, start=1):
            # Construct the complete file path
            source_file = os.path.join(source_folder, file_name)
            # Construct a new file name, such as "document_1.pdf"
            new_file_name = f'{name}_{str(l + 1)}.pdf'
            destination_file = os.path.join(destination_folder, new_file_name)
            # Move and rename the file
            shutil.move(source_file, destination_file)
            print(f'Signs：{name}_{l + 1}------Renamed Successfully')
        # Open the original PDF file
        input_pdf = PdfReader(r"Renamed Signs\{}_{}.pdf".format(name, str(l + 1)))
        # Create a PDF write object
        output_pdf = PdfWriter()
        # Add the first page to the output PDF
        output_pdf.add_page(input_pdf.pages[0])
        # Save the new PDF as a file
        with open(r"Renamed Signs\{}_{}.pdf".format(name, str(l + 1)), "wb") as output_file:
            output_pdf.write(output_file)
        print("The first page of the PDF is preserved and saved as a new file")
driver.quit()
